#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Corpus Builder - Speech Dataset XML Packaging

Packages XML metadata for speech datasets using the generateDS-generated API.

Usage:
    python build_corpus.py --config corpus_config.yaml
"""

import argparse
import openpyxl
import hashlib
import logging
import os
import re
import wave
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

import yaml
import corpus_api as api

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# Configuration
# =============================================================================

@dataclass
class CorpusConfig:
    """Initialize configuration for corpus metadata"""
    # Corpus-level attributes
    name: str = "unnamed_corpus"
    language: str = "unknown"
    source: str = "unknown"
    genre: str = "unknown"
    
    # Recording-level attributes (user-defined defaults)
    audio_format: str = "unknown"
    confidence_score: str = "-1"
    
    # Data source options
    xlsx_file: Optional[str] = None  # Option 1: Excel file
    audio_dir: str = "./audio"      # Option 2: Source directories
    text_dir: str = "./transcriptions"
    
    # Output
    output_file: str = "corpus.xml"
    csv_output_file: Optional[str] = None
    
    @classmethod
    def from_yaml(cls, yaml_path: str) -> "CorpusConfig":
        """Load configuration from a YAML file."""
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        return cls(**data)


@dataclass
class RecordingData:
    """Data container for a single recording."""
    audio_path: str
    filename: str
    speaker_id: str
    transcription: str = ""
    duration: str = "unknown"
    md5sum: str = "unknown"
    recording_date: str = "unknown"
    flag: str = ""


# =============================================================================
# Data Extraction Functions
# =============================================================================

def get_audio_duration(audio_path: str) -> str:
    """Extract duration from a WAV audio file."""
    try:
        with wave.open(audio_path, 'rb') as wav_file:
            frames = wav_file.getnframes()
            rate = wav_file.getframerate()
            duration = frames / float(rate)
            return f"{duration:.2f}"
    except Exception as e:
        logger.warning(f"Could not extract duration from {audio_path}: {e}")
        return "unknown"


def get_file_md5(file_path: str) -> str:
    """Calculate MD5 checksum of a file."""
    try:
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except Exception as e:
        logger.warning(f"Could not calculate MD5 for {file_path}: {e}")
        return "unknown"


def read_transcription(text_path: str) -> str:
    """Read transcription from a text file."""
    try:
        with open(text_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        logger.warning(f"Transcription file not found: {text_path}")
        return ""
    except Exception as e:
        logger.warning(f"Could not read transcription from {text_path}: {e}")
        return ""


def parse_recording_date(speaker_id: str) -> str:
    """Parse recording date from speaker ID (folder name).
    
    Converts format like '05h04jan2024' to '05h_04_jan_2024'
    """
    # Pattern: digits + 'h' + digits + month + year
    match = re.match(r'(\d+)h(\d+)([a-z]+)(\d+)', speaker_id, re.IGNORECASE)
    if match:
        hour, day, month, year = match.groups()
        return f"{hour}h_{day}_{month}_{year}"
    return speaker_id


def load_xlsx_metadata(xlsx_path: str) -> List[Dict[str, str]]:
    """Load metadata from an Excel (.xlsx) file.
    
    Expected columns: filename, duration, transcription
    Returns a list of recording metadata dictionaries.
    """
    records = []
    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheet = workbook.active
        
        # Get header row to find column indices
        headers = [cell.value.lower() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        filename_idx = headers.index('filename') if 'filename' in headers else None
        duration_idx = headers.index('duration') if 'duration' in headers else None
        transcription_idx = headers.index('transcription') if 'transcription' in headers else None
        flag_idx = headers.index('flag') if 'flag' in headers else None
        
        if filename_idx is None:
            logger.error("Excel file missing required 'filename' column")
            return records
        
        # Read data rows (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            filename = str(row[filename_idx]) if row[filename_idx] else ''
            if not filename:
                continue
            
            # Remove .wav extension if present for consistency
            base_filename = filename[:-4] if filename.endswith('.wav') else filename
            
            records.append({
                'filename': filename,
                'base_filename': base_filename,
                'duration': str(row[duration_idx]) if duration_idx is not None and row[duration_idx] else 'unknown',
                'transcription': str(row[transcription_idx]) if transcription_idx is not None and row[transcription_idx] else '',
                'flag': str(row[flag_idx]) if flag_idx is not None and row[flag_idx] is not None else ''
            })
        
        workbook.close()
        logger.info(f"Loaded {len(records)} entries from Excel file")
    except Exception as e:
        logger.error(f"Could not load Excel metadata from {xlsx_path}: {e}")
    
    return records


# =============================================================================
# Corpus Builder
# =============================================================================

class CorpusBuilder:
    """Main class for building speech corpus XML."""
    
    def __init__(self, config: CorpusConfig):
        self.config = config
        self.corpus: Optional[api.corpus] = None
        self.recordings: List[RecordingData] = []
        self.xlsx_records: List[Dict[str, str]] = []
        
        # Determine data source
        if config.xlsx_file:
            logger.info(f"Using Option 1: Loading metadata from Excel file")
            self.xlsx_records = load_xlsx_metadata(config.xlsx_file)
        else:
            logger.info(f"Using Option 2: Extracting metadata from source directories")
    
    def _extract_speaker_id_from_filename(self, filename: str) -> str:
        """Extract speaker ID from filename.
        
        Assumes filename format: <prefix>_<speaker_id>_<segment>.wav
        e.g., 'sama_nbl_05h04jan2024_0001.wav' -> '05h04jan2024'
        """
        # Try to extract speaker ID pattern (e.g., 05h04jan2024)
        match = re.search(r'(\d+h\d+[a-z]+\d+)', filename, re.IGNORECASE)
        if match:
            return match.group(1)
        
        # Fallback: use second-to-last underscore segment
        parts = filename.replace('.wav', '').split('_')
        if len(parts) >= 2:
            return parts[-2]
        
        return "unknown_speaker"
    
    @staticmethod
    def rename_to_sama_path(audio_path: str) -> str:
        """ 
        Rename path to use the final SAMA packaging path in the xml file, which is "sama/audio/<bulletin>/<audio_filename>"
        """
        p = Path(audio_path)
        bulletin = p.parent.name
        filename = p.name
        return f"sama/audio/{bulletin}/{filename}"

    def build_from_xlsx(self) -> List[RecordingData]:
        """Build recording list from Excel metadata (Option 1).
        
        Uses the xlsx file as the primary source for duration and transcription.
        Audio directory is still required for MD5 calculation.
        """
        recordings = []
        audio_dir = Path(self.config.audio_dir)
        
        if not audio_dir.exists():
            logger.warning(f"Audio directory does not exist: {audio_dir}. MD5 checksums will be 'unknown'.")
        
        for record in self.xlsx_records:
            filename = record['filename']
            base_filename = record['base_filename']
            speaker_id = self._extract_speaker_id_from_filename(filename)
            
            # Local audio path on disk: audio_dir/<speaker_id>/<filename>
            local_audio_path = audio_dir / speaker_id / filename
            
            # Calculate MD5 from the real local file, if it exists
            if local_audio_path.exists():
                md5sum = get_file_md5(str(local_audio_path))
            else:
                md5sum = "unknown"
                logger.debug(f"Audio file not found for MD5: {local_audio_path}")
            
            # Path to be written into the XML (SAMA packaging path)
            sama_audio_path = self.rename_to_sama_path(str(local_audio_path))
            
            recording = RecordingData(
                audio_path=sama_audio_path,
                filename=base_filename,
                speaker_id=speaker_id,
                transcription=record['transcription'],
                duration=record['duration'],
                md5sum=md5sum,
                recording_date=parse_recording_date(speaker_id),
                flag=record.get('flag', "")
            )
            recordings.append(recording)
        
        logger.info(f"Built {len(recordings)} recordings from Excel metadata")
        return recordings
    
    def build_from_directories(self) -> List[RecordingData]:
        """Build recording list from source directories (Option 2).
        
        Discovers audio files and extracts metadata from audio/text files.
        """
        recordings = []
        audio_dir = Path(self.config.audio_dir)
        
        if not audio_dir.exists():
            logger.error(f"Audio directory does not exist: {audio_dir}")
            return recordings
        
        # Iterate through speaker subdirectories
        for speaker_dir in sorted(audio_dir.iterdir()):
            if not speaker_dir.is_dir():
                continue
            
            speaker_id = speaker_dir.name
            logger.info(f"Processing speaker: {speaker_id}")
            
            # Find all audio files for this speaker
            audio_files = sorted(speaker_dir.glob("*.wav"))
            
            for audio_file in audio_files:
                filename = audio_file.stem
                
                # Extract duration from audio file
                duration = get_audio_duration(str(audio_file))
                
                # Read transcription from text file
                transcription = ""
                if self.config.text_dir:
                    text_path = os.path.join(self.config.text_dir, filename + ".txt")
                    transcription = read_transcription(text_path)
                
                # Calculate MD5 from actual audio file
                md5sum = get_file_md5(str(audio_file))
                
                # Path to be written into the XML (SAMA packaging path)
                sama_audio_path = self.rename_to_sama_path(str(audio_file))
                
                recording = RecordingData(
                    audio_path=sama_audio_path,
                    filename=filename,
                    speaker_id=speaker_id,
                    transcription=transcription,
                    duration=duration,
                    md5sum=md5sum,
                    recording_date=parse_recording_date(speaker_id)
                )
                recordings.append(recording)
        
        logger.info(f"Discovered {len(recordings)} audio files from directories")
        return recordings
    
    def build(self) -> api.corpus:
        """Build the corpus XML structure."""
        logger.info(f"Building corpus: {self.config.name}")
        
        # Choose data source based on configuration
        if self.config.xlsx_file and self.xlsx_records:
            # Option 1: Use Excel file as primary source
            self.recordings = self.build_from_xlsx()
        else:
            # Option 2: Use source directories
            self.recordings = self.build_from_directories()
        
        if not self.recordings:
            logger.warning("No recordings found!")
            self.corpus = api.corpus(
                name=self.config.name,
                language=self.config.language,
                source=self.config.source,
                genre=self.config.genre
            )
            return self.corpus
        
        # Group recordings by speaker
        speakers_dict: Dict[str, List[RecordingData]] = {}
        for recording in self.recordings:
            if recording.speaker_id not in speakers_dict:
                speakers_dict[recording.speaker_id] = []
            speakers_dict[recording.speaker_id].append(recording)
        
        # Create corpus structure
        self.corpus = api.corpus(
            name=self.config.name,
            language=self.config.language,
            source=self.config.source,
            genre=self.config.genre
        )
        
        # Create speaker elements
        for speaker_id, speaker_recordings in sorted(speakers_dict.items()):
            speaker = api.speakerType(id=speaker_id)
            
            # Sort recordings by filename so XML lists segments in numeric order
            sorted_recordings = sorted(speaker_recordings, key=lambda r: r.filename)

            # Add recordings to speaker
            for rec in sorted_recordings:
                recording_elem = api.recordingType(
                    audio=rec.audio_path,
                    md5sum=rec.md5sum,
                    duration=rec.duration,
                    confidence_score=self.config.confidence_score,
                    recording_date=rec.recording_date,
                    audio_format=self.config.audio_format,
                    transcript_word=rec.transcription
                )
                speaker.add_recording(recording_elem)
            
            self.corpus.add_speaker(speaker)
            logger.info(f"Added speaker '{speaker_id}' with {len(speaker_recordings)} recordings")
        
        logger.info(f"Corpus built with {len(speakers_dict)} speakers")
        return self.corpus
    
    def save(self, output_path: Optional[str] = None) -> None:
        """Save the corpus to an XML file."""
        if self.corpus is None:
            logger.error("Corpus not built yet. Call build() first.")
            return
        
        output_path = output_path or self.config.output_file
        
        logger.info(f"Saving corpus to {output_path}")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            self.corpus.export(f, 0, pretty_print=True)
        
        logger.info(f"Corpus saved successfully to {output_path}")

    def save_csv_metadata(self, csv_path: Optional[str] = None) -> None:
        """Save selected metadata to a CSV file (filename, duration, flag, transcription).

        This does not change the XML; it is an auxiliary export based on self.recordings.
        """
        if not self.recordings:
            logger.error("No recordings available to export. Call build() first.")
            return

        # Determine target CSV path
        if csv_path is None:
            if self.config.csv_output_file:
                csv_path = self.config.csv_output_file
            else:
                csv_path = str(Path(self.config.output_file).with_suffix('.csv'))

        logger.info(f"Saving metadata CSV to {csv_path}")

        try:
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['filename', 'duration', 'flag', 'transcription'])

                # Sort by filename so rows appear in numeric order (e.g., ..._0001.wav, ..._0002.wav)
                for rec in sorted(self.recordings, key=lambda r: r.filename):
                    writer.writerow([
                        f"{rec.filename}.wav",
                        rec.duration,
                        rec.flag,
                        rec.transcription,
                    ])
            logger.info(f"Metadata CSV saved successfully to {csv_path}")
        except Exception as e:
            logger.error(f"Failed to save metadata CSV to {csv_path}: {e}")


# =============================================================================
# Command-Line Interface
# =============================================================================

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate XML metadata for speech datasets",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
  python build_corpus.py --config corpus_config.yaml
        """
    )
    
    parser.add_argument(
        '--config', '-c',
        type=str,
        required=True,
        help='Path to YAML configuration file'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Load configuration and build corpus
    logger.info(f"Loading configuration from {args.config}")
    config = CorpusConfig.from_yaml(args.config)
    
    builder = CorpusBuilder(config)
    builder.build()
    builder.save()
    builder.save_csv_metadata()


if __name__ == "__main__":
    main()
